import openpyxl
import random as rd

rv = []
for i in range(0, 101):
    rv.append(i)

class make_file() :

    def form1(num):

        # 원본 FORM 주소 선언
        filename = "formFile_pp/참여인력_FORM1.xlsx"

        book = openpyxl.load_workbook(filename)
        sheet = book.worksheets[0]

        # TXT에서 불러올 영역 선언
        CALL = [sheet['C9'], sheet['E9'], sheet['C11'], sheet['E11'], sheet['G13'], sheet['E23'], sheet['E25']]
        FILE = ['참여인력_FORM1_성명(이름)', '참여인력_FORM1_소속(학교 소속)', '참여인력_FORM1_최종학력(학위)(학력)', '참여인력_FORM1_전공',
                '참여인력_FORM1_사업참여기간(날짜)', '참여인력_FORM1_참여기간(날짜1, 날짜2)', '참여인력_FORM1_참여기간(날짜1, 날짜2)']

        i = 0

        while i < 7:
            CALL[i].value = data.call(FILE[i])
            i = i + 1

        sheet['I11'].value = data.wper(sheet['C11'].value) + '년'  # 근무경력 생성
        sheet['J13'].value = str(rd.choice(rv)) + '%'  # 참여율 생성

        # 담당업무 생성

        wbox = data.work(3)
        sheet['C13'].value = wbox[0]
        sheet['H23'].value = wbox[1]
        sheet['H25'].value = wbox[2]

        # 사업명, 사업개요 생성

        bs = data.bs(2)
        sheet['A23'].value = bs[0][0]
        sheet['C23'].value = bs[0][1]
        sheet['A25'].value = bs[1][0]
        sheet['C25'].value = bs[1][1]

        # 파일 저장

        book.save('../../_inputData/raw/documents/pp_form1_%d.xlsx' % num)


    def form2(num):

        # 원본 FORM 주소 선언
        filename = "formFile_pp/참여인력_FORM2.xlsx"

        book = openpyxl.load_workbook(filename)
        sheet = book.worksheets[0]

        # TXT에서 불러올 영역 선언
        CALL = [sheet['B9'], sheet['E9'], sheet['B13'], sheet['E13'], sheet['B17'], sheet['E32'], sheet['E35']]
        FILE = ['경력증명서_FORM2_성명(이름)', '경력증명서_FORM2_주민등록번호(주민)', '경력증명서_FORM2_소속(회사)', '경력증명서_FORM2_학위(학력)',
                '경력증명서_FORM2_참여기간(기간1)', '경력증명서_FORM2_참여기간(기간2, 기간3)', '경력증명서_FORM2_참여기간(기간2, 기간3)']

        i = 0

        while i < 7:
            if i != 3:
                CALL[i].value = data.call(FILE[i])
                i = i + 1

            else:  # 나이에 따른 학력 수정
                CALL[i].value = data.call(FILE[i])

                if int(sheet['E9'].value[0:2]) > 50:
                    bir = '19' + sheet['E9'].value[0:2]
                else:
                    bir = '20' + sheet['E9'].value[0:2]

                age = 2019 - int(bir)

                while (sheet['E13'].value[-2:] == '석사') and age <= 27:
                    CALL[i].value = data.call(FILE[i])
                while (sheet['E13'].value[-2:] == '박사') and age < 31:
                    while (True):
                        CALL[i].value = data.call(FILE[i])
                        if sheet['E13'].value[-2:] != '석사':
                            break
                i = i + 1

        # 생년월일에 따른 근속년수 만들기 + 입사년월 + 해당분야 경력

        if age <= 24:
            sheet['H9'].value = '2020.01'
            sheet['J9'].value = '0 년'
            sheet['H13'].value = str(rd.randint(0, 3)) + ' 년'
        else:
            sheet['H9'].value = str(int(bir) + 24) + '.' + str(rd.randint(1, 12))
            sheet['J9'].value = str(age - 24) + ' 년'
            sheet['H13'].value = str(age - 24 + rd.randint(0, 3)) + ' 년'  # 분야 근무 경력은 근속년수 + 0~3



        sheet['J17'].value = str(rd.choice(rv)) + '%'  # 참여율 생성
        sheet['J13'].value = data.lic(0)  # 자격증 랜덤
        sheet['H21'].value = ''

        # 활용가능 기술 생성
        teno = rd.randint(1, 4)
        box = data.tec(teno)
        for i in range(0, teno):
            sheet['H21'].value = sheet['H21'].value + box[i]
            if i != teno - 1:
                sheet['H21'].value = sheet['H21'].value + ", "

        # 사업명, 담당업무

        bsi = data.bsi(2)
        sheet['B32'].value = bsi[0][0]
        sheet['G32'].value = bsi[0][1]
        sheet['B35'].value = bsi[1][0]
        sheet['G35'].value = bsi[1][1]

        book.save('../../_inputData/raw/documents/pp_form2_%d.xlsx' % num)

    def form3(num):

        # 원본 FORM 주소 선언
        filename = "formFile_pp/참여인력_FORM3.xlsx"

        book = openpyxl.load_workbook(filename)
        sheet = book.worksheets[0]

        # TXT에서 불러올 영역 선언
        CALL = [sheet['B9'], sheet['E9'], sheet['H9'], sheet['B11'], sheet['G15'], sheet['I24'], sheet['I26']]
        FILE = ['참여인력_FORM3_성명(이름)', '참여인력_FORM3_소속', '참여인력_FORM3_직책(직위)', '참여인력_FORM3_최종학력(학력)',
                '참여인력_FORM3_사업참여기간(기간)', '참여인력_FORM3_회사명1_회사명2', '참여인력_FORM3_회사명1_회사명2']

        i = 0

        while i < 7:
            CALL[i].value = data.call(FILE[i])
            i = i + 1

        sheet['J15'].value = str(rd.choice(rv)) + '%'  # 참여율 생성
        sheet['I13'].value = data.lic(0)  # 자격증 랜덤

        # 학력에 따른 생년월일 제조
        if sheet['B11'].value[-2:] == '석사':
            yy = rd.randint(68, 91)
        elif sheet['B11'].value[-2:] == '박사':
            yy = rd.randint(68, 86)
        else:
            yy = rd.randint(68, 96)

        mm = rd.randint(1, 12)

        if mm == 2:
            dd = rd.randint(1, 28)
        elif ((mm % 2 == 1) and (mm < 9)) or ((mm % 2 == 0) and (mm > 9)):
            dd = rd.randint(1, 31)
        else:
            dd = rd.randint(1, 30)

        if mm > 9 and dd > 9:
            birth = str(yy) + str(mm) + str(dd)
        elif mm > 9 and dd <= 9:
            birth = str(yy) + str(mm) + '0' + str(dd)
        elif mm <= 9 and dd > 9:
            birth = str(yy) + '0' + str(mm) + str(dd)
        elif mm <= 9 and dd <= 9:
            birth = str(yy) + '0' + str(mm) + '0' + str(dd)

        sheet['J9'].value = birth
        wper = rd.randint(0, 120 - yy - 20)
        sheet['I11'].value = str(wper) + '년'

        # 수행년도 제작

        y1 = int(sheet['G15'].value[0:4]) - 2
        sheet['F24'].value = y1
        sheet['F26'].value = y1 + 1

        # 사업명, 사업개요, 담당업무

        bsi = data.bsi(2)
        sheet['A24'].value = bsi[0][2]
        sheet['D24'].value = bsi[0][0]
        sheet['H24'].value = bsi[0][1]
        sheet['A26'].value = bsi[1][2]
        sheet['D26'].value = bsi[1][0]
        sheet['H26'].value = bsi[1][1]

        book.save('../../_inputData/raw/documents/pp_form3_%d.xlsx' % num)
class data():
    def call(data_type):  # TXT의 DATA를 불러오는 함수

        fn = data_type

        if 'FORM1' in fn:  # FROM 1의 DATA를 다룰 경우

            txt = open("dataFile_pp/참여인력_FORM1/" + fn + ".txt", "r", encoding='utf=8')

            if fn == '참여인력_FORM1_사업참여기간(날짜)':
                i = 5000  # TXT가 몇 줄 있는지 확인
            elif fn == '참여인력_FORM1_성명(이름)':
                i = 1040
            elif fn == '참여인력_FORM1_전공':
                i = 46107
            elif fn == '참여인력_FORM1_참여기간(날짜1, 날짜2)':
                i = 5000
            elif fn == '참여인력_FORM1_최종학력(학위)(학력)':
                i = 4
            elif fn == '참여인력_FORM1_학교':
                i = 30
            elif fn == '참여인력_FORM1_소속(학교 소속)':
                i = 4679

        elif 'FORM2' in fn:  # FROM 2의 DATA를 다룰 경우

            txt = open("dataFile_pp/참여인력_FORM2/" + fn + ".txt", "r", encoding='utf=8')

            if fn == '경력증명서_FORM2_소속(회사)':
                i = 5015
            elif fn == '경력증명서_FORM2_성명(이름)':
                i = 1040
            elif fn == '경력증명서_FORM2_주민등록번호(주민)':
                i = 88775
            elif fn == '경력증명서_FORM2_참여기간(기간1)':
                i = 5000
            elif fn == '경력증명서_FORM2_참여기간(기간2, 기간3)':
                i = 5000
            elif fn == '경력증명서_FORM2_학위(학력)':
                i = 5656

        elif 'FORM3' in fn:  # FROM 3의 DATA를 다룰 경우

            txt = open("dataFile_pp/참여인력_FORM3/" + fn + ".txt", "r", encoding='utf=8')

            if fn == '참여인력_FORM3_사업참여기간(기간)':
                i = 5000
            elif fn == '참여인력_FORM3_성명(이름)':
                i = 1040
            elif fn == '참여인력_FORM3_소속':
                i = 11
            elif fn == '참여인력_FORM3_직책(직위)':
                i = 60
            elif fn == '참여인력_FORM3_최종학력(학력)':
                i = 5656
            elif fn == '참여인력_FORM3_회사명1_회사명2':
                i = 5015

        j = 1
        k = rd.randint(1, i)  # 지정한 TXT중 랜덤한 줄 지정

        while j <= k:
            inList = txt.readline()  # 지정한 TXT의 랜덤한 줄 추출
            j = j + 1

        result = inList.strip()

        return result

    def wper(recent):  # 최종학력에 따른 분야 근무 경력 생성

        if recent == '고졸':
            per = rd.randint(0, 10)  # 고졸의 경우 근무경력 0~10년
        elif recent == '학사':
            per = rd.randint(0, 5)  # 학사의 경우 근무경력 0~5년
        elif recent == '석사':
            per = rd.randint(3, 8)  # 석사의 경우 근무경력 3~8년
        elif recent == '박사':
            per = rd.randint(6, 15)  # 박사의 경우 근무경력 6~15년

        return str(per)

    def lic(self):  # 자격증 랜덤 호출

        lic = ['없음', '정보처리기사', '워드프로세서 1급', '워드프로세서 2급', '워드프로세서 3급',
               '정보검색사 1급', '정보검색사 2급', '전문검색사', '시스템관리사', '정보설계사',
               'IPCT Special', 'IPCT General', '정보처리산업기사', '정보처리 기능사', 'PCT A',
               'PCT B', '컴퓨터활용능력 1급', '컴퓨터활용능력 2급', 'ITQ A', 'ITQ B',
               'PC 정비사 1급', 'PC 정비사 2급', '회계정보사 1급', '회계정보사 2급', '회계정보사 3급',
               '컴퓨터그래픽스 운용사', 'CNA', 'CNE', 'NCNE', 'CNI',
               'CIP', 'Oracle 7 DBA', 'Oracle 8 DBA', 'Oracle AD', 'SCJP',
               'SCJD', 'MCPS', 'MCSE', 'MCSD', 'MCT']

        i = rd.randint(0, 39)

        return str(lic[i])

    def tec(num):  # 사용 가능 언어 랜덤 생성

        tec = ['Python', 'JAVA', 'C', 'C++', 'C#',
               'VisualBasic', 'Ruby', 'PHP', 'SQL', 'HTML5',
               'Scala', 'Pascal', 'Swift', 'R', 'MATLAB']
        k = 14
        tlist = []
        for i in range(0, num):  # 언어 중복을 피하기 위해, 뽑은 언어는 ARRAY에서 덮어씌운다.
            j = rd.randint(0, k)
            tlist.append(tec[j])
            tec[j] = tec[k]
            k = k - 1

        return tlist

    def work(num):  # FROM 1의 경우, 전공이 너무 다양하여 세부적인 담당업무를 지정하기 어려우므로 만만한 업무들만 투입

        work = ['진행보조', '연구보조', '사무보조', '자료분석', '통계분석',
                '유지보수', '운영지원', '고객응대', '자료입력', '단순보조',
                '분류업무', '타자입력', '행정보조', '행정사무', '자료제작',
                '전산업무', '고객상담', '단순사무', '영상물제작', '업무지원',
                '근무지원', '행정', '조사보조', '조사', '관리보조',
                '현장지원', '행정지원', '사무지원', '자료수집', '정보검색']
        k = 29
        wlist = []
        for i in range(0, num):  # tec함수와 동일 기능
            j = rd.randint(0, k)
            wlist.append(work[j])
            work[j] = work[k]
            k = k - 1

        return wlist

    def bs(num):  # form1의 사업명과 사업내용은 1:1대응하며, 다중배열로 매칭한다.

        bsn = ['오스카', '신진푸드리서치 14회', '사성전자 s10', '교통량 분석', 'HIAS-1192',
               '한림 모터쇼', 'WCC', '전국 경제인 연합 컨퍼런스', 'TGS', '한국물리학학회 학술대회',
               'V6엔진 개발', '국가미래산업 투자 설명회', '암호화폐 컨퍼런스', '대학가자!캠프', '에듀넷 입시설명회',
               '건축사 실기 시험', '베이비페어', '국군의날 행사', '지킬 앤 하이드', '르네 마그리트전',
               'KBS 방송제', '아름다운 음악회', '법회 켐페인', '나눔푸드페어', '한국치위생협회 홈페이지 제작',
               '성남 뷰티 콘테스트', 'ONCE 5th Album', '드림캠퍼스', '내츄럴 피지크', '(주)Leebok festa',
               '이브루아', '대한비만치료학회', '부산 락 페스티벌', '힘나정', '학진사 교과서 편찬',
               '국방 R&D 예산분배회', '쉐프코리아 시즌 4', '(주)유니더스 대표캐릭터 선정', '3070', '부산 국제 게임 쇼',
               '아이리스', '진영 리크루트', 'PJS 라이브', '경기 아트 그랑프리', '<르 생의 정원>',
               '대한 암 학회', '공익광고협의회 흡연근절주간', '아이카 온라인 쇼케이스', '아이키퍼', 'KJ 국제 헤어 페스티벌']

        bsc = ['온라인 게임 개발', '식품광고제', '전자제품 출시행사', '분기 교통량 분석', '전자기기 신제품 개발',
               '(주)한림 모터쇼', '세계 컴퓨터 컨퍼런스', '전국 경제인 연합 컨퍼런스', '도쿄 게임쇼', '한국물리학학회 학술대회',
               '차세대 하이브리드 엔진 개발', '국가미래산업 19기(2020~2023) 투자 설명회', '암호화폐 컨퍼런스', '고교1년생 대상 학습동기부여 캠프', '17학년도 입시설명회',
               '금년도 건축사 실기 시험', '육아용품 페스티벌', '금년도 국군의날 행사', '지킬 앤 하이드 뮤지컬 내한 공연', '르네 마그리트 전시회',
               '금년도 한국방송 방송제', 'KBS 주관 음악회', '대한 불교 조계종 켐페인', '나눔식품 푸드 페스티벌', '한국치위생협회 홈페이지 제작',
               '화장품 브랜드 홍보', '그룹 ONCE 5번째 앨범 제작', '고교생 대상 진로캠프', '보디빌딩 대회', '스포츠웨어 신상품 홍보',
               '미용 화학제품 개발', '비만치료학회 세미나', '금년도 부산 락 페스티벌', '씨알리스 복제약 제작', '신규 교육과정 교과서 제작',
               '금년도 국방 R&D 예산분배회', 'N.NET 주관 조리 오디션 프로그램', '기업 이미지 캐릭터 제작', '제약회사 신제품 광고제작', '국제 게임쇼',
               '(주)페일게임즈 제작 모바일 게임', '(주)진영사 주관 신인작가 발굴 프로젝트', '아이돌그룹 콘서트', '경기 아트 그랑프리', '프랑스 유명 소설 수입관련 설명회',
               '대한 암 학회 정기 세미나', '금연 홍보물 제작', '신작 온라인게임 쇼케이스', 'PC 사용 제한 프로그램 개발', '한.일 헤어 페스티벌']

        k = 49

        totlist = []

        for i in range(0, num):  # tec함수와 동일 기능이지만, 2중배열로 return한다.
            j = rd.randint(0, k)
            totlist.append([bsn[j], bsc[j]])
            bsn[j] = bsn[k]
            bsc[j] = bsc[k]
            k = k - 1
        return totlist

    def bsi(num):  # form2,3은 담당업무가 개발로 고정되어있으므로, 사업명과 사업내용,담당업무를 다중배열로 매칭한다.

        bsin = ['한국식물병리학회 홈페이지 제작', '한국대학교 공학대학 홈페이지', 'MUJI 홈페이지 리뉴얼', '차량 관리 DB 구축', '한양대학교 학생 DBMS 개발',
                '미래일보 홈페이지 개편', '하림시네마 예매시스템 개발', '배달맨 모바일', '무신사 DB 구축', '조선호텔 고객관리 DB',
                '스포맥스 주문 시스템 개발', '한세항공 좌석지정 서비스', '한양대학교 소프트웨어융합대학 안내 페이지 개편', '하스 온라인 모바일 버젼 개발',
                '개인 학기 시간표 제작 서비스 개발',
                'SPOT TV 인터넷 개인 방송국', '해인 게스트하우스 숙박예약 시스템', '미세미세 모바일 어플리케이션', '사운드하운드 - 음악 추천 어플리케이션 개발',
                '웹툰 제공 서비스 DB 구축',
                '벌꿀툰 모바일 페이지 제작', '코엑스 컨퍼런스 홀 일정 관리 시스템', '코로나 예방 - 가까운 확진자 찾기 어플리케이션', '구골 이미지 검색 시스템',
                '온라인 쇼핑몰 DB 구축',
                'ROL 전적 검색 사이트 개발', '한국 맥도날드 가까운 매장찾기 서비스', '치킨의 제왕 어플리케이션 개발', '경리나라 사무 문서 제공 서비스 구축',
                '한국암치료학회 학회자료 DBMS 개발',
                '아스카온라인 OTP 시스템', 'GS손해보험 온라인 상담 시스템', '병무청 입영일자 본인선택 서비스', '씨네39 리뷰 관리 시스템 개발', '스토리 PC방 고객 DB 구축',
                '자인한방병원 통합예약시스템', '텔레그램 PC버젼', '대한모터스 K9 자율주행 시스템', '아만다 인연찾기 서비스 개발', '공협은행 원장 DBMS 개편',
                '식약일보 모바일 뉴스레터 개발', '티켓나라 통합 예매 시스템', '<해주세요> 심부름 어플 개발', '중고 자동차 거래 서비스 개발', '상록구청 주민 DB',
                '한양대학교 전자우편 시스템', '피닉스파크 콘도 예약 시스템', '코인넷 암호화폐 거래소 사이트 개편', 'YES25 도서 검색 서비스 개발', '한양 도서관 자료 검색 DB', ]

        bsit = ['한국식물병리학회 홈페이지', '한국대학교 공학대학 홈페이지', 'MUJI 홈페이지', '차량 관리 DB', '한양대학교 학생 DBMS',
                '미래일보 홈페이지', '하림시네마 예매시스템', '배달맨 모바일', '무신사 DB', '조선호텔 고객관리 DB',
                '스포맥스 주문 시스템', '한세항공 좌석지정 서비스', '한양대학교 소프트웨어융합대학', '하스 온라인', '개인 학기 시간표 제작 서비스',
                'SPOT TV', '해인 게스트하우스', '미세미세', '사운드하운드', '웹툰 제공 서비스 DB',
                '벌꿀툰 모바일 페이지', '코엑스 컨퍼런스', '코로나 예방 - 가까운 확진자 찾기 어플리케이션', '구골 이미지 검색 시스템', '온라인 쇼핑몰 DB 구축',
                'OP.GG', '맥딜리버리', '치킨의 제왕', '경리나라 사무 문서 제공 서비스', '한국암치료학회 학회자료 DB',
                '아스카온라인', 'GS손해보험 온라인 상담 시스템', '대한민국 병무청', '씨네39 리뷰 관리 시스템', '스토리 PC방 고객 DB',
                '자인한방병원 통합예약시스템', '텔레그램 PC버젼', '대한모터스', '아만다 - 아무도 만나지 않는다', '공협은행',
                '식약일보 모바일 뉴스레터', '티켓나라', '해주세요', '중고 자동차 거래 서비스 개발', '상록구청',
                '한양대학교', '피닉스파크', '코인넷', 'YES25 도서 검색 서비스', '한양 도서관', ]

        cont = ['백엔드 개발', '프론트 엔드 개발', '뷰 개발', '데이터 셋 작성', '쿼리 작성']

        k = 49

        totlist = []

        for i in range(0, num):
            j = rd.randint(0, k)
            totlist.append([bsin[j], cont[(j % 5)], bsit[j]])
            bsin[j] = bsin[k]
            bsit[j] = bsit[k]
            k = k - 1

        return totlist

# 작성할 FORM 번호 입력

form_no = input('원하는 폼 선택 (1,2,3 중 입력) : ')
num = int(form_no)

# 선언한 FORM의 개수 입력

q = input('type %d formFile_pp form 파일을 몇 개 만드시겠습니까? : ' %num)
qu = int(q)

if num == 1 :
    i = 1
    while i < qu + 1 :
        make_file.form1(i)
        i = i + 1

elif num == 2 :
    i = 1
    while i < qu + 1 :
        make_file.form2(i)
        i = i + 1

elif num == 3 :
    i = 1
    while i < qu + 1 :
        make_file.form3(i)
        i = i + 1

print('form %d 로 %d개 완성되었습니다' %(num ,qu))