import openpyxl
import random
from docx.enum.text import WD_ALIGN_PARAGRAPH
import docx

class create:
    def __init__(self, n):
        self.n = n

    #form1 만드는 함수    
    def form1(self):
    # 경력증명서_FORM1 이름 txt파일 불러오기, 이름,주소,기간,부서1과 2는 통합되있음
        f1 = open("dataFile_ldcc/이름.txt", "r", encoding="UTF-8")
        f2 = open("dataFile_ldcc/주소.txt", "r", encoding="UTF-8")
        f3 = open("dataFile_ldcc/회사명.txt", "r", encoding="UTF-8")
        f4 = open("dataFile_ldcc/직위.txt", "r", encoding="UTF-8")
        f5 = open("dataFile_ldcc/견적일자.txt", "r", encoding="UTF-8")
        f6 = open("dataFile_ldcc/월(개월).txt", "r", encoding="UTF-8")
        f7 = open("dataFile_ldcc/참여율.txt", "r", encoding="UTF-8")
        f8 = open("dataFile_ldcc/월급여.txt", "r", encoding="UTF-8")
        f9 = open("dataFile_ldcc/사업자번호.txt", "r", encoding="UTF-8")

        # 텍스트파일 줄별 데이터를 리스트로 저장(홍길동1, 홍길동2 ...)
        list1 = f1.readlines() #이름
        list2 = f2.readlines() #주소
        list3 = f3.readlines() #회사명
        list4 = f4.readlines()
        list5 = f5.readlines()
        list6 = f6.readlines()
        list7 = f7.readlines()
        list8 = f8.readlines()
        list9 = f9.readlines()
        # 리스트 줄바꿈 모두 제거
        list1_f = [line.strip() for line in list1]
        list2_f = [line.strip() for line in list2]
        list3_f = [line.strip() for line in list3]
        list4_f = [line.strip() for line in list4]
        list5_f = [line.strip() for line in list5]
        list6_f = [line.strip() for line in list6]
        list7_f = [line.strip() for line in list7]
        list8_f = [line.strip() for line in list8]
        list9_f = [line.strip() for line in list9]
        # 텍스트파일 줄 길이를 저장하여, 랜덤 리스트 생성에 쓰임
        len_f1 = len(list1)
        len_f2 = len(list2)
        len_f3 = len(list3)
        len_f4 = len(list4)
        len_f5 = len(list5)
        len_f6 = len(list6)
        len_f7 = len(list7)
        len_f8 = len(list8)
        len_f9 = len(list9)
        # 파일닫기
        f1.close()
        f2.close()
        f3.close()
        f4.close()
        f5.close()
        f6.close()
        f7.close()
        f8.close()
        f9.close()
        # 엑셀파일 불러오기
        file = openpyxl.load_workbook("formFile_ldcc/ldcc_form11.xlsx")
        # 엑셀파일 액티브해서 시트 불러오기
        sheet = file.active
        # 1번 폼에 이름넣는 함수 지정, 모든 리스트 갯수만큼 반복실행하고 행을 구분
        for k in range(1, self.n + 1):
            com1 = list3_f[random.randrange(len_f3)]
            com2 = list3_f[random.randrange(len_f3)]
            pos1 = list4_f[random.randrange(len_f4)]
            pos2 = list4_f[random.randrange(len_f4)]
            day = list6_f[random.randrange(len_f6)]
            jo1 = list7_f[random.randrange(len_f7)]
            jo2 = list7_f[random.randrange(len_f7)]
            mo1 = list8_f[random.randrange(len_f8)]
            mo2 = list8_f[random.randrange(len_f8)]
            mo3 = (int(mo1) + int(mo2) * 3) * (1 / 10)
            mo4 = int(mo1) + int(mo2) * 3
            mo5 = mo3 + mo4
            sheet['F7'].value = list4_f[random.randrange(len_f4)] + ' ' + list1_f[random.randrange(len_f1)] + ' (인)'
            sheet['A4'].value = '■ TO : ' + com1 + ' 귀하'
            sheet['F10'].value = '■ 납품 장소 : ' + com1
            sheet['F5'].value = '상호 : ' + com2
            sheet['F4'].value = '사업자 등록번호 : ' + list9_f[random.randrange(len_f9)]
            sheet['F6'].value = '주소 : ' + list2_f[random.randrange(len_f2)]
            sheet['A10'].value = '■ 견적 일자 : ' + list5_f[random.randrange(len_f5)]
            sheet['B16'].value = pos1
            sheet['B17'].value = pos2
            sheet['B18'].value = pos2
            sheet['B19'].value = pos2
            sheet['D16'].value = day
            sheet['D17'].value = day
            sheet['D18'].value = day
            sheet['D19'].value = day
            sheet['C16'].value = int(mo1)
            sheet['F16'].value = int(mo1)
            sheet['C17'].value = int(mo2)
            sheet['F17'].value = int(mo2)
            sheet['C18'].value = int(mo2)
            sheet['F18'].value = int(mo2)
            sheet['C19'].value = int(mo2)
            sheet['F19'].value = int(mo2)
            sheet['E16'].value = jo1
            sheet['E17'].value = jo2
            sheet['E18'].value = jo2
            sheet['E19'].value = jo2
            sheet['F24'].value = mo3
            sheet['F23'].value = mo4
            sheet['F28'].value = mo5
            sheet['F30'].value = mo5
            sheet['C9'].value = mo5
            if k <= 9:
                file.save("../../_inputData/raw/documents/ldcc_form1_0%s.xlsx" % k)
            else:
                file.save("../../_inputData/raw/documents/ldcc_form1_%s.xlsx" % k)

    def form2(self):
        com = open('datafile_ldcc/회사명.txt', 'r', encoding = 'UTF8')
        com_data = com.read()
        com = com_data.split('\n')

        date = open('datafile_ldcc/날짜2.txt', 'r', encoding = 'UTF8')
        date_data = date.read()
        date = date_data.split('\n')

        loc = open('datafile_ldcc/주소.txt', 'r', encoding = 'UTF8')
        loc_data = loc.read()
        loc = loc_data.split('\n')

        pos = open('datafile_ldcc/직위.txt', 'r', encoding = 'UTF8')
        pos_data = pos.read()
        pos = pos_data.split('\n')

        per = open('datafile_ldcc/이름.txt', 'r', encoding = 'UTF8')
        per_data = per.read()
        per = per_data.split('\n')

        file = docx.Document("formFile_ldcc/table.docx")

        for k in range(1, self.n + 1):
            com_index = com[random.randrange(len(com))]
            com_index2 = com[random.randrange(len(com))]
            date_inedex = date[random.randrange(len(date))]
            loc_index = loc[random.randrange(len(loc))]
            loc_index2 = loc[random.randrange(len(loc))]
            pos_index = pos[random.randrange(len(pos))]
            pos_index2 = pos[random.randrange(len(pos))]
            per_index = per[random.randrange(len(per))]
            per_index2 = per[random.randrange(len(per))]

            table = file.tables[0]
            table.cell(0, 0).text = str('저작권 합의서')
            table.cell(2, 0).text = str('%s(이하 "%s"이라 함와 "%s"는 "NER 고도화를 활용한 개인민감정보 검출 딥러닝 모델 개발" '
                                        '관련 %s 체결한 연구용역 계약서 (이하 "본 계약"이라함) 와 관련해 다음 내용에 합의한다.\n\n'
                                        '- 다음 -\n\n'
                                        '1. “%s”과 “%s”는 본 계약 산출물의 저작권을 공동으로 소유한다.\n'
                                        '2. "%s"과 "%s" 간 저작권의 지분율은 동등하다.\n'
                                        '3. "%s"은 "%s"의 동의 없이 자유롭게 저작권을 자기실시(저작권의 사용/수익/개작/변형/배포 및 '
                                        '제3자에게 영업/판매하여 수익을 얻는 행위를 포함한다. 이하 같다.)할 수 있다.\n'
                                        '4. "%s"이 저작권을 자기실시하여 덩는 수익 전부는 "%s"에 귀속된다.\n'
                                        '5. "%s"이 저작권을 자기실시가 아닌, 제3자에게 처분, 양도하여 수익을 얻는 경우, 해당 수익은 '
                                        '지분대로 배분된다. (단, 제반 비용은 제외한다)\n'
                                        '6. "%s"는 교육ㆍ학술 또는 연구를 위해 이용하는 경우를 제외하고 자기실시하지 않는다. 더불어 제3자와 '
                                        '본 계약의 산출물을 기반으로 교육ㆍ학술 또는 연구를 진행하는 경우, 사전에 “%s”의 동의를 받아야 한다.\n\n'
                                        '7. 본 합의서상에 규정되지 아니한 사항은 본 계약의 내용을 준용하며, 본 합의서의 규정이 본 계약의 내용과 상충될 경우 본 합의서를 우선하여 적용한다\n\n'
                                        '양 당사자는 이상 상호 합의된 내용이 자신의 의사와 일치함을 확인하며, 합의서 2부를 작성하여 기명날인하고 “%s”과 “%s”가 각 1부씩 보관한다.\n\n\n'
                                        '%s'
                                        %(com_index, com_index, com_index2, date_inedex, com_index, com_index2, com_index, com_index2, com_index, com_index2, com_index, com_index, com_index, com_index2, com_index, com_index, com_index2, date_inedex))
            table.cell(3, 0).text = str('%s\n'
                                        '%s\n'
                                        '%s %s (인)'
                                        %(com_index, loc_index, pos_index, per_index))
            table.cell(3, 1).text = str('%s\n'
                                        '%s\n'
                                        '%s %s (인)'
                                        %(com_index2, loc_index2, pos_index2, per_index2))
            if k <= 9:
                file.save("../../_inputData/raw/documents/ldcc_form2_0%s.docx" % k)
            else:
                file.save("../../_inputData/raw/documents/ldcc_form2_%s.docx" % k)