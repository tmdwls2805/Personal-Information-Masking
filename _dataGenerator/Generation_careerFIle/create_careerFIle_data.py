import openpyxl
import random

class create:
    def __init__(self, n):
        self.n = n

    #form1 만드는 함수    
    def form1(self):
    # 경력증명서_FORM1 이름 txt파일 불러오기, 이름,주소,기간,부서1과 2는 통합되있음
        f1 = open("dataFile_career/경력증명서_FORM1_성명(이름).txt", "r", encoding="UTF-8")
        f2 = open("dataFile_career/경력증명서_FORM1_전화번호(유선).txt", "r", encoding="UTF-8")
        f3 = open("dataFile_career/경력증명서_FORM1_(날짜).txt", "r", encoding="UTF-8")
        f4 = open("dataFile_career/경력증명서_FORM1_근무기간(기간1, 기간2).txt", "r", encoding="UTF-8")
        f5 = open("dataFile_career/경력증명서_FORM1_소속(학과).txt", "r", encoding="UTF-8")
        f6 = open("dataFile_career/경력증명서_FORM1_업체명(회사명).txt", "r", encoding="UTF-8")
        f7 = open("dataFile_career/경력증명서_FORM1_전화번호(핸드폰).txt", "r", encoding="UTF-8")
        f8 = open("dataFile_career/경력증명서_FORM1_주민번호(주민).txt", "r", encoding="UTF-8")
        f9 = open("dataFile_career/경력증명서_FORM1_주소.txt", "r", encoding="UTF-8")
        f10 = open("dataFile_career/경력증명서_FORM1_직위(직위).txt", "r", encoding="UTF-8")
        f11 = open("dataFile_career/경력증명서_FORM1_학력(학력).txt", "r", encoding="UTF-8")
        f12 = open("dataFile_career/경력증명서_FORM1_(날짜).txt", "r", encoding="UTF-8")
        f13 = open("dataFile_career/경력증명서_FORM1_근무부서(부서).txt", "r", encoding="UTF-8")
        # 텍스트파일 줄별 데이터를 리스트로 저장(홍길동1, 홍길동2 ...)
        list1 = f1.readlines()
        list2 = f2.readlines()
        list3 = f3.readlines()
        list4 = f4.readlines()
        list5 = f5.readlines()
        list6 = f6.readlines()
        list7 = f7.readlines()
        list8 = f8.readlines()
        list9 = f9.readlines()
        list10 = f10.readlines()
        list11 = f11.readlines()
        list12 = f13.readlines()
        # 리스트 줄바꿈 모두 제거
        list1_f = [line.strip() for line in list1]
        list2_f = [line.strip() for line in list2]
        list3_f = [line.strip() for line in list3]
        list4_f = [line.strip() for line in list4]
        list5_f = [line.strip() for line in list5]
        list6_f = [line.strip() for line in list6]
        list7_f = [line.strip() for line in list7]
        list8_f = [line.strip() for line in list8]
        list9_f = [line.strip() for line in list9]
        list10_f = [line.strip() for line in list10]
        list11_f = [line.strip() for line in list11]
        list12_f = [line.strip() for line in list12]
        # 회사명 list중에 (주),㈜,주식회사 들어간거 전부 ""로 replace
        new_list6_f = [s.replace('주식회사', '').replace('(주)','').replace('㈜','') for s in list6_f]
    # 텍스트파일 줄 길이를 저장하여, 랜덤 리스트 생성에 쓰임
        len_f1 =len(list1)
        len_f2 =len(list2)
        len_f3 =len(list3)
        len_f4 =len(list4)
        len_f5 =len(list5)
        len_f6 =len(list6)
        len_f7 =len(list7)
        len_f8 =len(list8)
        len_f9 =len(list9)
        len_f10 =len(list10)
        len_f11 =len(list11)
        len_f12 =len(list12)
        #파일닫기
        f1.close()
        f2.close()
        f3.close()
        f4.close()
        f5.close()
        f6.close()
        f7.close()
        f8.close()
        f9.close()
        f10.close()
        f11.close()
        f12.close()
        #엑셀파일 불러오기
        file = openpyxl.load_workbook("formFile_career/경력증명서_FORM1.xlsx")
        #엑셀파일 액티브해서 시트 불러오기
        sheet = file.active
        #1번 폼에 이름넣는 함수 지정, 모든 리스트 갯수만큼 반복실행하고 행을 구분
        for k in range(1, self.n +1):
            sheet['C9'].value = list1_f[random.randrange(len_f1)]
            sheet['C11'].value = list11_f[random.randrange(len_f11)]
            sheet['H11'].value = list5_f[random.randrange(len_f5)]
            sheet['C13'].value = list8_f[random.randrange(len_f8)]
            sheet['H13'].value = list7_f[random.randrange(len_f7)]
            sheet['C15'].value = list9_f[random.randrange(len_f9)]
            sheet['C17'].value = new_list6_f[random.randrange(len_f6)]
            sheet['H19'].value = list2_f[random.randrange(len_f2)]
            sheet['B25'].value = list4_f[random.randrange(len_f4)]
            sheet['D33'].value = list10_f[random.randrange(len_f10)]
            sheet['H33'].value = list1_f[random.randrange(len_f1)] #이름2
            sheet['C21'].value = list9_f[random.randrange(len_f9)] #주소2
            sheet['B27'].value = list4_f[random.randrange(len_f4)] #기간2
            sheet['F25'].value = list12_f[random.randrange(len_f12)]
            sheet['F27'].value = list12_f[random.randrange(len_f12)]
            sheet['A37'].value = "위와 같이 경력을 증명합니다.\n\n" + list3_f[random.randrange(len_f3)] + "\n\n주식회사 %s 대표 %s    (인)"% (sheet['C17'].value, list1_f[random.randrange(len_f1)])
            if k <= 9:
                file.save("../../_inputData/raw/documents/career_form1_0%s.xlsx" % k)
            else:
                file.save("../../_inputData/raw/documents/career_form1_%s.xlsx" % k)

    #form2 만드는 함수
    def form2(self):
    # 경력증명서_FORM2 이름 txt파일 불러오기, 주소는 통합되있음
        f1 = open("dataFile_career/경력증명서_FORM2_성명(이름).txt", "r", encoding="UTF-8")
        f2 = open("dataFile_career/경력증명서_FORM2_생년월일.txt", "r", encoding="UTF-8")
        f3 = open("dataFile_career/경력증명서_FORM2_주소(주소1), 소재지(주소2).txt", "r", encoding="UTF-8")
        f4 = open("dataFile_career/경력증명서_FORM2_회사명(회사명1)_(회사명1).txt", "r", encoding="UTF-8")
        f5 = open("dataFile_career/경력증명서_FORM2_직위(직위).txt", "r", encoding="UTF-8")
        f6 = open("dataFile_career/경력증명서_FORM2_재직기간(날짜1).txt", "r", encoding="UTF-8")
        f7 = open("dataFile_career/경력증명서_FORM2_(날짜2).txt", "r", encoding="UTF-8")
        f8 = open("dataFile_career/경력증명서_FORM2_근무부서(부서).txt", "r", encoding="UTF-8")
        # 텍스트파일 줄별 데이터를 리스트로 저장(홍길동1, 홍길동2 ...)
        list1 = f1.readlines()
        list2 = f2.readlines()
        list3 = f3.readlines()
        list4 = f4.readlines()
        list5 = f5.readlines()
        list6 = f6.readlines()
        list7 = f7.readlines()
        list8 = f8.readlines()
        #리스트 줄바꿈 모두 제거
        list1_f = [line.strip() for line in list1]
        list2_f = [line.strip() for line in list2]
        list3_f = [line.strip() for line in list3]
        list4_f = [line.strip() for line in list4]
        list5_f = [line.strip() for line in list5]
        list6_f = [line.strip() for line in list6]
        list7_f = [line.strip() for line in list7]
        list8_f = [line.strip() for line in list8]
        # 회사명 list중에 (주),㈜,주식회사 들어간거 전부 ""로 replace
        new_list4_f = [s.replace('주식회사', '').replace('(주)', '').replace('㈜', '') for s in list4_f]

        # 텍스트파일 줄 길이를 저장하여, 랜덤 리스트 생성에 쓰임
        len_f1 =len(list1)
        len_f2 =len(list2)
        len_f3 =len(list3)
        len_f4 =len(list4)
        len_f5 =len(list5)
        len_f6 =len(list6)
        len_f7 =len(list7)
        len_f8 =len(list8)
        #파일닫기
        f1.close()
        f2.close()
        f3.close()
        f4.close()
        f5.close()
        f6.close()
        f7.close()
        f8.close()
        #엑셀파일 불러오기
        file = openpyxl.load_workbook("formFile_career/경력증명서_FORM2.xlsx")
        #엑셀파일 액티브해서 시트 불러오기
        sheet = file.active
        #1번 폼에 이름넣는 함수 지정, 모든 리스트 갯수만큼 반복실행하고 행을 구분
        for k in range(1, self.n +1):
            sheet['C5'].value = list1_f[random.randrange(len_f1)]
            sheet['H5'].value = list2_f[random.randrange(len_f2)]
            sheet['C7'].value = list3_f[random.randrange(len_f3)]
            sheet['C9'].value = new_list4_f[random.randrange(len_f4)]
            sheet['C11'].value = list3_f[random.randrange(len_f3)]
            sheet['C13'].value = list8_f[random.randrange(len_f8)]
            sheet['H13'].value = list5_f[random.randrange(len_f5)]
            sheet['C17'].value = list6_f[random.randrange(len_f6)]
            sheet['A19'].value = "    상기와 같이 경력사항을 증명합니다.\n    (   용도 :  계약    )\n\n    %s\n\n  주식회사 %s   (인)" %(list7_f[random.randrange(len_f7)], sheet['C9'].value)
            if k <= 9:
                file.save("../../_inputData/raw/documents/career_form2_0%s.xlsx" % k)
            else:
                file.save("../../_inputData/raw/documents/career_form2_%s.xlsx" % k)
    #form3 만드는 함수
    def form3(self):
    # 경력증명서_FORM3 이름 txt파일 불러오기, 이름1,2,3 직위1,2 주소1,2는 통합되있음
        f1 = open("dataFile_career/경력증명서_FORM3_성명(이름1), 대표자(이름2), 담당자(이름3).txt", "r", encoding="UTF-8")
        f2 = open("dataFile_career/경력증명서_FORM3_주소(주소1), 소재지(주소2).txt", "r", encoding="UTF-8")
        f3 = open("dataFile_career/경력증명서_FORM3_직위(직위1)_직책(직위2).txt", "r", encoding="UTF-8")
        f4 = open("dataFile_career/경력증명서_FORM3_주민등록번호(주민).txt", "r", encoding="UTF-8")
        f5 = open("dataFile_career/경력증명서_FORM3_재직기간(기간).txt", "r", encoding="UTF-8")
        f6 = open("dataFile_career/경력증명서_FORM3_연락처(유선).txt", "r", encoding="UTF-8")
        f7 = open("dataFile_career/경력증명서_FORM2_회사명(회사명1)_(회사명1).txt", "r", encoding="UTF-8")
        f8 = open("dataFile_career/경력증명서_FORM3_(날짜).txt", "r", encoding="UTF-8")
        f9 = open("dataFile_career/경력증명서_FORM3_근무부서(부서1)_담당부서(부서2).txt", "r", encoding="UTF-8")
    # 텍스트파일 줄별 데이터를 리스트로 저장(홍길동1, 홍길동2 ...)
        list1 = f1.readlines()
        list2 = f2.readlines()
        list3 = f3.readlines()
        list4 = f4.readlines()
        list5 = f5.readlines()
        list6 = f6.readlines()
        list7 = f7.readlines()
        list8 = f8.readlines()
        list9 = f9.readlines()
    # 텍스트파일 줄바꿈 리스트 모두 제거
        list1_f = [line.strip() for line in list1]
        list2_f = [line.strip() for line in list2]
        list3_f = [line.strip() for line in list3]
        list4_f = [line.strip() for line in list4]
        list5_f = [line.strip() for line in list5]
        list6_f = [line.strip() for line in list6]
        list7_f = [line.strip() for line in list7]
        list8_f = [line.strip() for line in list8]
        list9_f = [line.strip() for line in list9]
        # 회사명 list중에 (주),㈜,주식회사 들어간거 전부 ""로 replace
        new_list7_f = [s.replace('주식회사', '').replace('(주)', '').replace('㈜', '') for s in list7_f]
    # 텍스트파일 줄 길이를 저장하여, 랜덤 리스트 생성에 쓰임
        len_f1 =len(list1)
        len_f2 =len(list2)
        len_f3 =len(list3)
        len_f4 =len(list4)
        len_f5 =len(list5)
        len_f6 =len(list6)
        len_f7 =len(list7)
        len_f8 = len(list8)
        len_f9 = len(list9)

        #파일닫기
        f1.close()
        f2.close()
        f3.close()
        f4.close()
        f5.close()
        f6.close()
        f7.close()
        f8.close()
        f9.close()
        #엑셀파일 불러오기
        file = openpyxl.load_workbook("formFile_career/경력증명서_FORM3.xlsx")
        #엑셀파일 액티브해서 시트 불러오기
        sheet = file.active
        #1번 폼에 이름넣는 함수 지정, 모든 리스트 갯수만큼 반복실행하고 행을 구분
        for k in range(1, self.n +1):
            sheet['D9'].value = list1_f[random.randrange(len_f1)] #이름1
            sheet['D15'].value = list1_f[random.randrange(len_f1)] #이름2
            sheet['H29'].value = list1_f[random.randrange(len_f1)] #이름3
            sheet['D11'].value = list2_f[random.randrange(len_f2)] #주소1
            sheet['D17'].value = list2_f[random.randrange(len_f2)] #주소2
            sheet['H19'].value = list3_f[random.randrange(len_f3)] #직위1
            sheet['D31'].value = list3_f[random.randrange(len_f3)] #직위2
            sheet['H9'].value = list4_f[random.randrange(len_f4)] #주민
            sheet['D23'].value = list5_f[random.randrange(len_f5)] #재직기간
            sheet['H31'].value = list6_f[random.randrange(len_f6)] #유선연락처
            sheet['D13'].value = new_list7_f[random.randrange(len_f7)] #회사명
            sheet['D19'].value = list9_f[random.randrange(len_f9)] #부서1 임의
            sheet['D29'].value = list9_f[random.randrange(len_f9)] #부서2 임의
            sheet['A33'].value = "위와 같이 경력을 증명합니다.\n\n" + list8_f[random.randrange(len_f8)] + "\n\n주식회사 %s 대표 %s   (인)" % (sheet['D13'].value, sheet['D15'].value)
            if k <= 9:
                file.save("../../_inputData/raw/documents/career_form3_0%s.xlsx" % k)
            else:
                file.save("../../_inputData/raw/documents/career_form3_%s.xlsx" % k)
